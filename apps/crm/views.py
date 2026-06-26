import csv
import io

from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.parsers import FormParser, JSONParser, MultiPartParser
from rest_framework.permissions import IsAuthenticated
from rest_framework import status

from apps.core.exceptions import APIResponse
from apps.core.permissions import IsOrganizationMember
from apps.crm.models import Activity, Contact, ContactGroup, Lead, PipelineStage
from apps.crm.serializers import (
    ActivitySerializer,
    ContactGroupSerializer,
    ContactSerializer,
    LeadSerializer,
    PipelineStageSerializer,
)


class TenantViewSetMixin:
    permission_classes = [IsAuthenticated, IsOrganizationMember]

    def perform_create(self, serializer):
        serializer.save(organization=self.request.organization)


class ContactViewSet(TenantViewSetMixin, viewsets.ModelViewSet):
    serializer_class = ContactSerializer
    parser_classes = [JSONParser, MultiPartParser, FormParser]
    filterset_fields = ["source", "assigned_to", "is_active"]
    search_fields = ["first_name", "last_name", "phone", "email", "company"]
    ordering_fields = ["created_at", "last_contacted_at"]

    def get_queryset(self):
        qs = Contact.objects.select_related("assigned_to").prefetch_related("groups").all()
        group_id = self.request.query_params.get("group")
        if group_id:
            qs = qs.filter(groups__id=group_id)
        return qs

    def create(self, request, *args, **kwargs):
        data = request.data.copy()
        phone = self._normalize_phone(data.get("phone", ""))
        if not phone:
            return APIResponse.error("Phone number is required.", status_code=400)

        data["phone"] = phone
        contact = Contact.objects.filter(organization=request.organization, phone=phone).first()
        if contact:
            serializer = self.get_serializer(contact, data=data, partial=True)
            serializer.is_valid(raise_exception=True)
            contact = serializer.save()
            return APIResponse.success(
                ContactSerializer(contact, context=self.get_serializer_context()).data,
                message="Contact updated",
            )

        serializer = self.get_serializer(data=data)
        serializer.is_valid(raise_exception=True)
        serializer.save(organization=request.organization)
        return APIResponse.success(
            serializer.data,
            message="Contact created",
            status_code=status.HTTP_201_CREATED,
        )

    @action(detail=False, methods=["post"], url_path="import")
    def import_contacts(self, request):
        upload = request.FILES.get("file")
        group_id = request.data.get("group_id")
        if not upload:
            return APIResponse.error("Upload a CSV or XLSX file.", status_code=400)

        group = None
        if group_id:
            group = ContactGroup.objects.filter(organization=request.organization, id=group_id).first()
            if not group:
                return APIResponse.error("Contact group not found.", status_code=404)

        try:
            rows = self._read_contact_rows(upload)
        except ValueError as exc:
            return APIResponse.error(str(exc), status_code=400)

        created = updated = skipped = 0
        errors = []
        for index, row in enumerate(rows, start=2):
            phone = self._normalize_phone(row.get("phone") or row.get("mobile") or row.get("number") or "")
            if not phone:
                skipped += 1
                errors.append({"row": index, "error": "Missing phone"})
                continue
            defaults = {
                "first_name": row.get("first_name") or row.get("name") or "",
                "last_name": row.get("last_name") or "",
                "email": row.get("email") or "",
                "company": row.get("company") or "",
                "source": Contact.Source.WHATSAPP,
                "is_active": True,
            }
            contact, was_created = Contact.objects.update_or_create(
                organization=request.organization,
                phone=phone,
                defaults=defaults,
            )
            if group:
                group.contacts.add(contact)
            created += 1 if was_created else 0
            updated += 0 if was_created else 1

        return APIResponse.success({
            "created": created,
            "updated": updated,
            "skipped": skipped,
            "errors": errors[:25],
        }, message="Contacts imported")

    def _read_contact_rows(self, upload):
        name = upload.name.lower()
        if name.endswith(".csv"):
            text = upload.read().decode("utf-8-sig")
            return list(csv.DictReader(io.StringIO(text)))
        if name.endswith(".xlsx"):
            try:
                from openpyxl import load_workbook
            except ImportError as exc:
                raise ValueError("Install openpyxl to import Excel files.") from exc
            wb = load_workbook(upload, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return []
            headers = [str(h or "").strip().lower() for h in rows[0]]
            return [
                {headers[i]: value for i, value in enumerate(row) if i < len(headers)}
                for row in rows[1:]
            ]
        raise ValueError("Unsupported file type. Use CSV or XLSX.")

    def _normalize_phone(self, value):
        digits = "".join(ch for ch in str(value) if ch.isdigit())
        if len(digits) == 10:
            return f"91{digits}"
        if digits.startswith("00"):
            return digits[2:]
        return digits


class ContactGroupViewSet(TenantViewSetMixin, viewsets.ModelViewSet):
    serializer_class = ContactGroupSerializer
    search_fields = ["name", "category"]
    filterset_fields = ["category", "is_active"]

    def get_queryset(self):
        return ContactGroup.objects.prefetch_related("contacts").all()


class PipelineStageViewSet(TenantViewSetMixin, viewsets.ModelViewSet):
    serializer_class = PipelineStageSerializer

    def get_queryset(self):
        return PipelineStage.objects.all()


class LeadViewSet(TenantViewSetMixin, viewsets.ModelViewSet):
    serializer_class = LeadSerializer
    filterset_fields = ["stage", "priority", "assigned_to", "is_archived"]
    search_fields = ["title", "contact__first_name", "contact__phone"]
    ordering_fields = ["created_at", "score", "value"]

    def get_queryset(self):
        return Lead.objects.select_related("contact", "stage", "assigned_to").all()

    @action(detail=True, methods=["post"])
    def move_stage(self, request, pk=None):
        lead = self.get_object()
        stage_id = request.data.get("stage_id")
        stage = PipelineStage.objects.filter(id=stage_id).first()
        if not stage:
            return APIResponse.error("Invalid stage", status_code=400)
        lead.stage = stage
        lead.save(update_fields=["stage", "updated_at"])
        return APIResponse.success(LeadSerializer(lead).data, message="Lead moved")


class ActivityViewSet(TenantViewSetMixin, viewsets.ModelViewSet):
    serializer_class = ActivitySerializer
    filterset_fields = ["type", "is_completed", "lead", "contact"]
    ordering_fields = ["due_at", "created_at"]

    def get_queryset(self):
        return Activity.objects.select_related("lead", "contact", "assigned_to").all()

    def perform_create(self, serializer):
        serializer.save(
            organization=self.request.organization,
            created_by=self.request.user,
        )
